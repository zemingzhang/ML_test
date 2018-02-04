from __future__ import division
from openpyxl import load_workbook
import os, sys
import numpy
from math import radians, cos, sin, asin, sqrt
from datetime import datetime
from openpyxl import Workbook
from sklearn.cluster import DBSCAN
import matplotlib.pyplot as plt
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)



print("sys.argv = ",len(sys.argv))

if len(sys.argv) < 2:
   print("please input filename")
   sys.exit("sorry, goodbye!");

openfilename=sys.argv[1]
eps_in=0.001



if len(sys.argv) == 3:
   eps_in=sys.argv[2]

wb = load_workbook(openfilename,read_only=True)
print(wb.sheetnames)
sheetnames = wb.get_sheet_names()
sheet = wb.get_sheet_by_name(sheetnames[0])
 
print "Work Sheet Titile:" ,sheet.title  
print sheet.max_row
#print sheet.max_column


print("toll:",sheet.max_row+1)

content = []
for row in sheet.rows:
   temp_list = []
   for cell in row:
      temp_list.append(cell.value)
   content.append(temp_list)
   #curX=temp_list[1]
   #curY=temp_list[2]
   
#print(content)
X=numpy.array(content)
#plt.scatter(X[:, 0], X[:, 1])
#plt.show()

#print(X)
y_pred = DBSCAN(eps = eps_in,min_samples = 85).fit_predict(content)

print("Max is ",numpy.max(y_pred))
plt.scatter(X[:, 0], X[:, 1], c=y_pred)
plt.show()


# write xlsm file
new_wb = Workbook(write_only=True)
new_sheet = new_wb.create_sheet()
offset=0
if (len(content)+1) > 400000:
    offset=2
if (len(content)+1) > 800000:
    offset=3
#for row in xrange(1, (len(content)+1)):

if offset == 0:
   for row in xrange(1, len(content)+1):
       new_sheet.append([content[row-1][0],content[row-1][1],y_pred[row-1]])

   new_wb.save(filename='new_file.xlsx')

if offset == 2:
   for row in xrange(1, 400000):
    #for col in range(1, 5):
       
       #new_sheet.cell(row=row, column=col).value = content[row-1][col-1]
       new_sheet.append([content[row-1][0],content[row-1][1],content[row-1][2],content[row-1][3]])
    #new_sheet.append([content[row-1][0],content[row-1][1],content[row-1][2]])
    #print(content[row-1][col-1])
   new_wb.save(filename='new_file.xlsx')

   new_wb1 = Workbook(write_only=True)
   new_sheet1 = new_wb1.create_sheet()
   for row in xrange(400000, len(content)+1):
       new_sheet1.append([content[row-1][0],content[row-1][1],content[row-1][2],content[row-1][3]])

   new_wb1.save(filename='new_file1.xlsx')
   
   
   
   
if offset == 3:
   for row in xrange(1, 400000):
       new_sheet.append([content[row-1][0],content[row-1][1],content[row-1][2],content[row-1][3]])

   new_wb.save(filename='new_file.xlsx')

   new_wb1 = Workbook(write_only=True)
   new_sheet1 = new_wb1.create_sheet()
   for row in xrange(400000, 800000):
       new_sheet1.append([content[row-1][0],content[row-1][1],content[row-1][2],content[row-1][3]])

   new_wb1.save(filename='new_file1.xlsx')
   
   new_wb2 = Workbook(write_only=True)
   new_sheet2 = new_wb2.create_sheet()
   for row in xrange(800000, len(content)+1):
       new_sheet2.append([content[row-1][0],content[row-1][1],content[row-1][2],content[row-1][3]])

   new_wb2.save(filename='new_file2.xlsx')